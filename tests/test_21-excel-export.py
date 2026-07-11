import pytest
import io
import database.db
from openpyxl import load_workbook


@pytest.fixture
def client(monkeypatch, tmp_path):
    db_path = str(tmp_path / "test.db")
    monkeypatch.setattr(database.db, "DB_PATH", db_path)

    # Import app inside fixture so monkeypatching takes effect
    from app import app as flask_app

    flask_app.config.update({"TESTING": True, "WTF_CSRF_ENABLED": False})

    database.db.init_db(db_path)
    database.db.seed_db(db_path)

    with flask_app.test_client() as client:
        yield client


def test_export_excel_auth_guard(client):
    """
    Test that unauthenticated requests to export Excel redirect to login.
    """
    response = client.get("/expenses/export/excel")
    assert response.status_code == 302
    assert "login" in response.headers["Location"]


def test_export_excel_metadata_and_headers(client):
    """
    Test that exporting Excel returns correct mimetype, headers, and filename.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})
    response = client.get("/expenses/export/excel")
    assert response.status_code == 200

    expected_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.mimetype == expected_mime

    disposition = response.headers.get("Content-Disposition", "")
    assert "attachment" in disposition
    assert 'filename="spendly_expenses_' in disposition
    assert disposition.endswith('.xlsx"')


def test_export_excel_sheet_name(client):
    """
    Test that the downloaded Excel workbook has a sheet named 'Expenses'.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})
    response = client.get("/expenses/export/excel")
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.data))
    assert "Expenses" in wb.sheetnames


def test_export_excel_headers(client):
    """
    Test that the Excel workbook contains correct column headers in the first row.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})
    response = client.get("/expenses/export/excel")
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.data))
    ws = wb["Expenses"]

    header_row = [ws.cell(row=1, column=col).value for col in range(1, 5)]
    assert header_row == ["Date", "Category", "Description", "Amount"]


def test_export_excel_content(client):
    """
    Test that the Excel workbook contains correct data rows from seeded expenses.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})
    response = client.get("/expenses/export/excel")
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.data))
    ws = wb["Expenses"]

    # Default seed_db inserts 8 expenses for demo user
    # Row 1 = header, rows 2..9 = data
    assert ws.max_row == 9  # 1 header + 8 data rows

    # Verify amount values are stored as numeric (not strings)
    for row_idx in range(2, ws.max_row + 1):
        amount_val = ws.cell(row=row_idx, column=4).value
        assert isinstance(amount_val, (int, float))
        assert amount_val > 0


def test_export_excel_with_filtering(client):
    """
    Test that active filters are respected in the Excel output.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})

    # Filter by category Food (original seed has exactly 2 Food expenses)
    response = client.get("/expenses/export/excel?category=Food")
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.data))
    ws = wb["Expenses"]

    # Should have header + 2 Food expenses
    assert ws.max_row == 3

    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row=row_idx, column=2).value == "Food"


def test_export_excel_empty_results(client):
    """
    Test that export returns a valid Excel workbook with just headers when no matches.
    """
    client.post("/login", data={"email": "demo@spendly.com", "password": "demo123"})

    # Filter for non-existent category/date combination
    response = client.get("/expenses/export/excel?category=Bills&start_date=2099-01-01")
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.data))
    ws = wb["Expenses"]

    # Only the header row should be present
    assert ws.max_row == 1
    header_row = [ws.cell(row=1, column=col).value for col in range(1, 5)]
    assert header_row == ["Date", "Category", "Description", "Amount"]
