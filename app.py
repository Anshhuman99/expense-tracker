from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    g,
    get_flashed_messages,
    abort,
    Response,
)
import json
from database.db import (
    get_db,
    init_db,
    seed_db,
    create_user,
    get_user_by_email,
    create_expense,
    update_expense,
    delete_expense as db_delete_expense,
)
from database.queries import (
    get_user_by_id,
    get_summary_stats,
    get_recent_transactions,
    get_category_breakdown,
    get_categories,
    get_filtered_expenses,
    get_filtered_expenses_count,
    get_expense_by_id,
    create_budget,
    get_budget_by_id,
    get_budgets_for_month,
    update_budget,
    delete_budget as db_delete_budget,
    get_month_category_spending,
    get_monthly_spending_trend,
    get_category_spending_breakdown,
    get_highest_expense,
)
from werkzeug.security import check_password_hash
import sqlite3
import re
import datetime
import math
import calendar
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
ALLOWED_CATEGORIES = [
    "Food",
    "Transport",
    "Bills",
    "Health",
    "Entertainment",
    "Shopping",
    "Other",
]
EXPENSES_PER_PAGE = 10


def _get_filter_params():
    """
    Parse filter parameters from request query arguments.
    """
    category = request.args.get("category", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    raw_search = request.args.get("search_query") or request.args.get("q") or ""
    search_query = raw_search.strip()[:100]
    sort_by = request.args.get("sort_by", "date").strip()
    order = request.args.get("order", "DESC").strip()
    return {
        "category": category,
        "start_date": start_date,
        "end_date": end_date,
        "search_query": search_query,
        "sort_by": sort_by,
        "order": order,
    }


def _sanitize_csv_value(val):
    """
    Sanitize values to prevent CSV injection (Formula Injection).
    If a string value starts with =, +, -, or @, prepend a single quote '.
    """
    if not val:
        return ""
    str_val = str(val)
    if str_val and str_val[0] in ("=", "+", "-", "@"):
        return f"'{str_val}"
    return str_val


# Brand colour used in exported Excel header rows (matches --color-primary in CSS)
EXCEL_HEADER_COLOR = "2D6A4F"


def _build_expense_workbook(expenses):
    """
    Build and return an openpyxl Workbook for the given expenses list.
    Produces a styled header row, data rows with numeric amounts, and
    approximate column widths. Sanitises text cells to prevent formula
    injection when the workbook is opened in spreadsheet software.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=EXCEL_HEADER_COLOR,
        end_color=EXCEL_HEADER_COLOR,
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Date", "Category", "Description", "Amount"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, exp in enumerate(expenses, start=2):
        ws.cell(row=row_idx, column=1, value=exp["date"])
        ws.cell(row=row_idx, column=2, value=_sanitize_csv_value(exp["category"]))
        ws.cell(row=row_idx, column=3, value=_sanitize_csv_value(exp["description"]))
        amount_cell = ws.cell(
            row=row_idx, column=4, value=round(float(exp["amount"]), 2)
        )
        amount_cell.number_format = "#,##0.00"

    col_widths = [12, 14, 40, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


app = Flask(__name__)
app.secret_key = "dev-secret-key"

with app.app_context():
    init_db()
    seed_db()


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    if user_id is None:
        g.user = None
    else:
        conn = get_db()
        g.user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        conn.close()


# ------------------------------------------------------------------ #
# Routes                                                              #
# ------------------------------------------------------------------ #


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if session.get("user_id"):
        return redirect(url_for("profile"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not name:
            flash("Name is required.", "error")
            return render_template("register.html")

        if not email or not EMAIL_REGEX.match(email):
            flash("Invalid email address.", "error")
            return render_template("register.html")

        if len(password) < 8:
            flash("Password must be at least 8 characters long.", "error")
            return render_template("register.html")

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("register.html")

        try:
            create_user(name, email, password)
            flash("Account created! Please log in.", "success")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("Email already registered.", "error")
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("profile"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        user = get_user_by_email(email)
        if not user or not check_password_hash(user["password_hash"], password):
            flash("Invalid email or password.", "error")
            return render_template("login.html")

        session.clear()
        session["user_id"] = user["id"]
        session["user_name"] = user["name"]
        flash("Logged in successfully.", "success")
        return redirect(url_for("profile"))

    return render_template("login.html")


@app.route("/demo")
def demo_login():
    import random
    from werkzeug.security import generate_password_hash

    email = "demo@spendly.com"
    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if not user:
            # Create user if not exists
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO users (name, email, password_hash) VALUES (?, ?, ?)",
                ("Demo User", email, generate_password_hash("demo123")),
            )
            conn.commit()
            user = conn.execute(
                "SELECT * FROM users WHERE email = ?", (email,)
            ).fetchone()

        user_id = user["id"]

        # Clear existing expenses and budgets to avoid infinite growth and keep it clean
        conn.execute("DELETE FROM expenses WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM budgets WHERE user_id = ?", (user_id,))

        # Seeding parameters
        num_expenses = 1200
        today = datetime.date.today()

        # Descriptions dictionary
        desc_map = {
            "Food": [
                "Swiggy order",
                "Zomato lunch",
                "Grocery shopping",
                "Chai and snacks",
                "Dinner with friends",
                "Cafe coffee",
                "Breakfast at Udipi",
                "Fruit market",
            ],
            "Transport": [
                "Uber ride",
                "Ola cab",
                "Metro recharge",
                "Auto fare",
                "Petrol refill",
                "Train ticket",
            ],
            "Bills": [
                "Electricity bill",
                "Broadband internet",
                "Mobile recharge",
                "Water bill",
                "Gas cylinder refill",
                "Rent payment",
                "DTH subscription",
            ],
            "Shopping": [
                "Amazon purchase",
                "Myntra clothes",
                "Shoes from Nike",
                "Supermarket run",
                "Electronics item",
                "Bookstore purchase",
            ],
            "Entertainment": [
                "Netflix subscription",
                "Spotify Premium",
                "Movie ticket",
                "Weekend concert",
                "Gaming arcade",
                "Bowling",
                "Theme park ticket",
            ],
            "Health": [
                "Pharmacy medicines",
                "Doctor consultation",
                "Lab test",
                "Health checkup",
                "Gym membership",
                "Eye exam",
            ],
            "Other": [
                "Courier charges",
                "Laundry",
                "Gift for colleague",
                "Miscellaneous cash spending",
                "Photocopy charges",
            ],
        }

        categories_weight = [
            "Food",
            "Transport",
            "Bills",
            "Shopping",
            "Entertainment",
            "Health",
            "Other",
        ]
        # Probability weights for each category
        weights = [0.35, 0.20, 0.10, 0.15, 0.10, 0.05, 0.05]

        expense_records = []
        for _ in range(num_expenses):
            category = random.choices(categories_weight, weights=weights, k=1)[0]

            # Select amount range based on category
            if category == "Food":
                amount = round(random.uniform(50, 800), 2)
            elif category == "Transport":
                amount = round(random.uniform(20, 500), 2)
            elif category == "Bills":
                amount = round(random.uniform(200, 3000), 2)
            elif category == "Shopping":
                amount = round(random.uniform(200, 5000), 2)
            elif category == "Entertainment":
                amount = round(random.uniform(100, 1500), 2)
            elif category == "Health":
                amount = round(random.uniform(100, 2000), 2)
            else:  # Other
                amount = round(random.uniform(50, 1000), 2)

            # Random date in the past 180 days (6 months)
            days_ago = random.randint(0, 180)
            expense_date = (today - datetime.timedelta(days=days_ago)).strftime(
                "%Y-%m-%d"
            )

            description = random.choice(desc_map[category])
            expense_records.append(
                (user_id, amount, category, expense_date, description)
            )

        # Insert all in one batch transaction
        conn.executemany(
            "INSERT INTO expenses (user_id, amount, category, date, description) VALUES (?, ?, ?, ?, ?)",
            expense_records,
        )

        # Let's seed budgets for the past 6 months + current month.
        # Get list of unique YYYY-MM values in the date range
        months_to_seed = set()
        for i in range(185):
            m_str = (today - datetime.timedelta(days=i)).strftime("%Y-%m")
            months_to_seed.add(m_str)

        budget_records = []
        # Category budgets
        cat_budgets = {
            "Food": 15000.0,
            "Transport": 5000.0,
            "Bills": 12000.0,
            "Shopping": 10000.0,
            "Entertainment": 6000.0,
            "Health": 4000.0,
            "Other": 3000.0,
        }
        for month in months_to_seed:
            for cat, amt in cat_budgets.items():
                budget_records.append((user_id, cat, amt, month))

        conn.executemany(
            "INSERT INTO budgets (user_id, category, amount, month) VALUES (?, ?, ?, ?)",
            budget_records,
        )

        conn.commit()

        # Log in
        session.clear()
        session["user_id"] = user["id"]
        session["user_name"] = user["name"]
        flash(
            "Logged into Demo Account with a pre-populated dataset of past 6 months (1200 expenses)!",
            "success",
        )
        return redirect(url_for("profile"))

    except Exception as e:
        conn.rollback()
        flash(f"Error initializing demo account: {str(e)}", "error")
        return redirect(url_for("landing"))
    finally:
        conn.close()


# ------------------------------------------------------------------ #
# Placeholder routes — students will implement these                  #
# ------------------------------------------------------------------ #


@app.route("/terms")
def terms():
    return render_template("terms.html")


@app.route("/privacy")
def privacy():
    return render_template("privacy.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "success")
    return redirect(url_for("landing"))


@app.route("/profile")
def profile():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    user = get_user_by_id(user_id)
    if not user:
        session.clear()
        flash("User session invalid. Please log in again.", "error")
        return redirect(url_for("login"))

    # Get active filters from URL query parameters
    active_filters = _get_filter_params()
    category = active_filters["category"]
    start_date = active_filters["start_date"]
    end_date = active_filters["end_date"]
    search_query = active_filters["search_query"]
    sort_by = active_filters["sort_by"]
    order = active_filters["order"]

    is_filtered = any([category, start_date, end_date, search_query])

    # Retrieve user's logged categories dynamically
    categories = get_categories(user_id)

    # Get live summary stats from database
    summary = get_summary_stats(user_id)

    # Calculate current month's spent amount dynamically
    current_month = datetime.date.today().strftime("%Y-%m")
    conn = get_db()
    month_spent_row = conn.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id = ? AND date LIKE ?",
        (user_id, f"{current_month}%"),
    ).fetchone()
    month_spent = month_spent_row[0] if month_spent_row[0] is not None else 0.0
    conn.close()

    stats = {
        "total_spent": summary["total_spent"],
        "month_spent": month_spent,
        "total_count": summary["transaction_count"],
        "top_category": summary["top_category"],
    }

    # Retrieve real category breakdown from database
    db_breakdown = get_category_breakdown(user_id)
    breakdown = [
        {"category": item["name"], "total": item["amount"], "percentage": item["pct"]}
        for item in db_breakdown
    ]

    # Retrieve real expenses (paginated, 10 per page)
    total_count = get_filtered_expenses_count(
        user_id=user_id,
        category=category,
        start_date=start_date,
        end_date=end_date,
        search_query=search_query,
    )

    per_page = EXPENSES_PER_PAGE
    total_pages = max(1, math.ceil(total_count / per_page))

    try:
        page = int(request.args.get("page", 1))
    except (ValueError, TypeError):
        page = 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    recent_expenses = get_filtered_expenses(
        user_id=user_id,
        category=category,
        start_date=start_date,
        end_date=end_date,
        search_query=search_query,
        sort_by=sort_by,
        order=order,
        limit=per_page,
        offset=offset,
    )

    start_idx = (page - 1) * per_page + 1 if total_count > 0 else 0
    end_idx = min(page * per_page, total_count)

    return render_template(
        "profile.html",
        user=user,
        stats=stats,
        breakdown=breakdown,
        recent_expenses=recent_expenses,
        categories=categories,
        active_filters=active_filters,
        is_filtered=is_filtered,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
        start_idx=start_idx,
        end_idx=end_idx,
    )


@app.route("/analytics")
def analytics():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    # Parse filter parameters
    category = request.args.get("category", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    # Validate date formats if provided
    if start_date:
        try:
            datetime.datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            flash("Invalid start date format. Expected YYYY-MM-DD.", "error")
            start_date = ""
    if end_date:
        try:
            datetime.datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            flash("Invalid end date format. Expected YYYY-MM-DD.", "error")
            end_date = ""

    active_filters = {
        "category": category,
        "start_date": start_date,
        "end_date": end_date,
    }
    is_filtered = any([category, start_date, end_date])

    # Categories for dropdown
    categories = get_categories(user_id)

    # Fetch filtered expenses to check if overall data exists (empty states)
    filtered_expenses = get_filtered_expenses(
        user_id=user_id,
        category=category,
        start_date=start_date,
        end_date=end_date,
    )

    # Use the helper function to query aggregated category totals
    breakdown_data = get_category_spending_breakdown(
        user_id=user_id,
        category=category if category else None,
        start_date=start_date if start_date else None,
        end_date=end_date if end_date else None,
    )

    # Monthly spending trend
    trend_data = get_monthly_spending_trend(
        user_id=user_id,
        category=category if category else None,
        start_date=start_date if start_date else None,
        end_date=end_date if end_date else None,
    )

    # Budget vs actual: use start_date month or current month
    if start_date:
        try:
            budget_month = datetime.datetime.strptime(start_date, "%Y-%m-%d").strftime(
                "%Y-%m"
            )
        except ValueError:
            budget_month = datetime.date.today().strftime("%Y-%m")
    else:
        budget_month = datetime.date.today().strftime("%Y-%m")

    budgets_for_month = get_budgets_for_month(user_id, budget_month)
    actual_spending = get_month_category_spending(user_id, budget_month)
    budget_comparison = [
        {
            "category": b["category"],
            "budget": round(b["amount"], 2),
            "actual": actual_spending.get(b["category"], 0.0),
        }
        for b in budgets_for_month
    ]

    return render_template(
        "analytics.html",
        categories=categories,
        active_filters=active_filters,
        is_filtered=is_filtered,
        has_data=len(filtered_expenses) > 0,
        breakdown_data=breakdown_data,
        trend_data=trend_data,
        budget_comparison=budget_comparison,
        has_budgets=len(budgets_for_month) > 0,
        budget_month=budget_month,
    )


@app.route("/insights")
def insights():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    # Fetch user info
    user = get_user_by_id(user_id)

    # Fetch all expenses
    all_expenses = get_filtered_expenses(user_id)
    has_data = len(all_expenses) > 0

    if not has_data:
        return render_template(
            "insights.html",
            user=user,
            has_data=False,
            stats={},
            insights_list=[],
            highest_expense=None,
        )

    # Calculate Lifetime Stats
    lifetime_total = sum(exp["amount"] for exp in all_expenses)
    transaction_count = len(all_expenses)

    # Group expenses by YYYY-MM
    monthly_expenses = {}
    for exp in all_expenses:
        m = exp["date"][:7]
        monthly_expenses.setdefault(m, []).append(exp)

    # Monthly average
    months_count = len(monthly_expenses)
    monthly_average = lifetime_total / months_count if months_count > 0 else 0.0

    # Current month's daily run rate
    today = datetime.date.today()
    current_month_str = today.strftime("%Y-%m")
    current_month_expenses = monthly_expenses.get(current_month_str, [])
    current_month_total = sum(exp["amount"] for exp in current_month_expenses)

    # Days elapsed in current month (or all days if month is in the past, but we focus on current month)
    days_in_month = today.day
    daily_average_current = (
        current_month_total / days_in_month if days_in_month > 0 else 0.0
    )

    # Project month-end total
    # Find number of days in the current month
    _, total_days_in_month = calendar.monthrange(today.year, today.month)
    projected_total = daily_average_current * total_days_in_month

    # MoM comparison
    # Get previous month string
    prev_month_date = today.replace(day=1) - datetime.timedelta(days=1)
    prev_month_str = prev_month_date.strftime("%Y-%m")
    prev_month_expenses = monthly_expenses.get(prev_month_str, [])
    prev_month_total = sum(exp["amount"] for exp in prev_month_expenses)

    mom_change_pct = 0.0
    if prev_month_total > 0:
        mom_change_pct = (
            (current_month_total - prev_month_total) / prev_month_total
        ) * 100

    # Top category breakdown percentage logic
    category_totals = {}
    for exp in all_expenses:
        category_totals[exp["category"]] = (
            category_totals.get(exp["category"], 0.0) + exp["amount"]
        )

    top_category = (
        max(category_totals, key=category_totals.get) if category_totals else None
    )
    top_category_pct = 0.0
    if top_category and lifetime_total > 0:
        top_category_pct = (category_totals[top_category] / lifetime_total) * 100

    # Highest single expense
    highest_expense = get_highest_expense(user_id)

    # Generate insights cards
    insights_list = []

    # 1. Month-over-month insight
    if prev_month_total > 0:
        if mom_change_pct > 10:
            insights_list.append(
                {
                    "type": "warning",
                    "icon": "trending_up",
                    "title": "Spending Surge",
                    "message": f"Your spending this month is {mom_change_pct:.1f}% higher than last month (₹{current_month_total:,.2f} vs ₹{prev_month_total:,.2f}).",
                }
            )
        elif mom_change_pct < -10:
            insights_list.append(
                {
                    "type": "success",
                    "icon": "trending_down",
                    "title": "Great Savings!",
                    "message": f"Your spending this month is {abs(mom_change_pct):.1f}% lower than last month (₹{current_month_total:,.2f} vs ₹{prev_month_total:,.2f}).",
                }
            )
        else:
            insights_list.append(
                {
                    "type": "info",
                    "icon": "trending_flat",
                    "title": "Stable Spending",
                    "message": f"Your spending this month is stable compared to last month (₹{current_month_total:,.2f} vs ₹{prev_month_total:,.2f}).",
                }
            )

    # 2. Top heavy warning
    if top_category and top_category_pct > 50:
        insights_list.append(
            {
                "type": "warning",
                "icon": "pie_chart",
                "title": f"Dominant Category: {top_category}",
                "message": f"{top_category} accounts for {top_category_pct:.1f}% of your total lifetime spending. Consider creating a monthly budget for this category.",
            }
        )

    # 3. High expense tip
    if highest_expense:
        insights_list.append(
            {
                "type": "info",
                "icon": "payments",
                "title": "Largest Single Expense",
                "message": f"Your purchase of '{highest_expense['description'] or highest_expense['category']}' on {highest_expense['date']} for ₹{highest_expense['amount']:,.2f} was your largest single expense.",
            }
        )

    # 4. Budget Warning
    # Fetch budgets for this month
    budgets_for_month = get_budgets_for_month(user_id, current_month_str)
    actual_spending = get_month_category_spending(user_id, current_month_str)
    for b in budgets_for_month:
        actual = actual_spending.get(b["category"], 0.0)
        limit = b["amount"]
        if actual > limit:
            insights_list.append(
                {
                    "type": "danger",
                    "icon": "error",
                    "title": f"Budget Exceeded: {b['category']}",
                    "message": f"You spent ₹{actual:,.2f} on {b['category']}, exceeding your budget of ₹{limit:,.2f} by ₹{(actual - limit):,.2f}.",
                }
            )
        elif actual > limit * 0.8:
            insights_list.append(
                {
                    "type": "warning",
                    "icon": "warning",
                    "title": f"Budget Alert: {b['category']}",
                    "message": f"You spent ₹{actual:,.2f} on {b['category']}, which is { (actual/limit)*100:.1f}% of your budget limit (₹{limit:,.2f}).",
                }
            )

    stats = {
        "lifetime_total": round(lifetime_total, 2),
        "transaction_count": transaction_count,
        "monthly_average": round(monthly_average, 2),
        "daily_average_current": round(daily_average_current, 2),
        "projected_total": round(projected_total, 2),
        "current_month_total": round(current_month_total, 2),
        "prev_month_total": round(prev_month_total, 2),
        "mom_change_pct": round(mom_change_pct, 2),
    }

    return render_template(
        "insights.html",
        user=user,
        has_data=True,
        stats=stats,
        insights_list=insights_list,
        highest_expense=highest_expense,
    )


@app.route("/expenses/export/csv")
def export_csv():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    active_filters = _get_filter_params()
    category = active_filters["category"]
    start_date = active_filters["start_date"]
    end_date = active_filters["end_date"]
    search_query = active_filters["search_query"]
    sort_by = active_filters["sort_by"]
    order = active_filters["order"]

    expenses = get_filtered_expenses(
        user_id=user_id,
        category=category,
        start_date=start_date,
        end_date=end_date,
        search_query=search_query,
        sort_by=sort_by,
        order=order,
        limit=None,
        offset=None,
    )

    csv_buffer = StringIO()
    csv_writer = csv.writer(csv_buffer)
    csv_writer.writerow(["Date", "Category", "Description", "Amount"])

    for exp in expenses:
        csv_writer.writerow(
            [
                exp["date"],
                _sanitize_csv_value(exp["category"]),
                _sanitize_csv_value(exp["description"]),
                f"{exp['amount']:.2f}",
            ]
        )

    timestamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"spendly_expenses_{timestamp}.csv"

    response = Response(csv_buffer.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.route("/expenses/export/excel")
def export_excel():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    active_filters = _get_filter_params()
    expenses = get_filtered_expenses(
        user_id=user_id, limit=None, offset=None, **active_filters
    )

    wb = _build_expense_workbook(expenses)
    excel_buffer = BytesIO()
    wb.save(excel_buffer)

    timestamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"spendly_expenses_{timestamp}.xlsx"

    response = Response(
        excel_buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.route("/expenses/add", methods=["GET", "POST"])
def add_expense():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        amount_str = request.form.get("amount", "").strip()
        category = request.form.get("category", "").strip()
        date_str = request.form.get("date", "").strip()
        description = request.form.get("description", "").strip()

        has_error = False
        amount = None

        if not amount_str:
            flash("Amount is required.", "error")
            has_error = True
        else:
            try:
                amount = float(amount_str)
                if not math.isfinite(amount) or amount <= 0:
                    flash("Amount must be a positive number.", "error")
                    has_error = True
            except ValueError:
                flash("Amount must be a valid number.", "error")
                has_error = True

        if not category:
            flash("Category is required.", "error")
            has_error = True
        elif category not in ALLOWED_CATEGORIES:
            flash("Invalid category selected.", "error")
            has_error = True

        if not date_str:
            flash("Date is required.", "error")
            has_error = True
        else:
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                flash("Invalid date format. Use YYYY-MM-DD.", "error")
                has_error = True

        if len(description) > 250:
            flash("Description cannot exceed 250 characters.", "error")
            has_error = True

        if has_error:
            return render_template(
                "add_expense.html",
                amount=amount_str,
                category=category,
                date=date_str,
                description=description,
                categories=ALLOWED_CATEGORIES,
            )

        create_expense(user_id, amount, category, date_str, description)
        flash("Expense added successfully!", "success")
        return redirect(url_for("profile"))

    # GET request
    default_date = datetime.date.today().strftime("%Y-%m-%d")
    return render_template(
        "add_expense.html",
        amount="",
        category="",
        date=default_date,
        description="",
        categories=ALLOWED_CATEGORIES,
    )


@app.route("/expenses/<int:id>/edit", methods=["GET", "POST"])
def edit_expense(id):
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    user = get_user_by_id(user_id)
    if not user:
        session.clear()
        flash("User session invalid. Please log in again.", "error")
        return redirect(url_for("login"))

    # Fetch the expense
    expense = get_expense_by_id(id)
    if not expense:
        abort(404)

    # Ownership check
    if expense["user_id"] != user_id:
        abort(403)

    if request.method == "POST":
        amount_str = request.form.get("amount", "").strip()
        category = request.form.get("category", "").strip()
        date_str = request.form.get("date", "").strip()
        description = request.form.get("description", "").strip()

        has_error = False
        amount = None

        if not amount_str:
            flash("Amount is required.", "error")
            has_error = True
        else:
            try:
                amount = float(amount_str)
                if not math.isfinite(amount) or amount <= 0:
                    flash("Amount must be a positive number.", "error")
                    has_error = True
            except ValueError:
                flash("Amount must be a valid number.", "error")
                has_error = True

        if not category:
            flash("Category is required.", "error")
            has_error = True
        elif category not in ALLOWED_CATEGORIES:
            flash("Invalid category selected.", "error")
            has_error = True

        if not date_str:
            flash("Date is required.", "error")
            has_error = True
        else:
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                flash("Invalid date format. Use YYYY-MM-DD.", "error")
                has_error = True

        if len(description) > 250:
            flash("Description cannot exceed 250 characters.", "error")
            has_error = True

        if has_error:
            return render_template(
                "edit_expense.html",
                expense=expense,
                amount=amount_str,
                category=category,
                date=date_str,
                description=description,
                categories=ALLOWED_CATEGORIES,
            )

        update_expense(id, amount, category, date_str, description)
        flash("Expense updated successfully!", "success")
        return redirect(url_for("profile"))

    # GET request: Load existing values
    return render_template(
        "edit_expense.html",
        expense=expense,
        amount=expense["amount"],
        category=expense["category"],
        date=expense["date"],
        description=expense["description"],
        categories=ALLOWED_CATEGORIES,
    )


@app.route("/expenses/<int:id>/delete", methods=["GET", "POST"])
def delete_expense(id):
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    user = get_user_by_id(user_id)
    if not user:
        session.clear()
        flash("User session invalid. Please log in again.", "error")
        return redirect(url_for("login"))

    expense = get_expense_by_id(id)
    if not expense:
        abort(404)

    if expense["user_id"] != user_id:
        abort(403)

    if request.method == "POST":
        db_delete_expense(id)
        flash("Expense deleted successfully!", "success")
        return redirect(url_for("profile"))

    # GET request
    return render_template(
        "delete_expense.html",
        expense=expense,
    )


# ------------------------------------------------------------------ #
# Budget Routes & Helpers                                            #
# ------------------------------------------------------------------ #


def validate_budget_data(category, amount_str, month_str):
    """
    Validates category, amount, and month for budget creation/updates.
    Returns (amount, error_messages_list).
    """
    errors = []
    amount = None

    if not category or category not in ALLOWED_CATEGORIES:
        errors.append("Please select a valid category.")

    if not amount_str:
        errors.append("Amount is required.")
    else:
        try:
            amount = float(amount_str)
            if not math.isfinite(amount) or amount <= 0:
                errors.append("Amount must be a positive number.")
        except ValueError:
            errors.append("Amount must be a valid number.")

    if not month_str:
        errors.append("Month is required.")
    else:
        try:
            datetime.datetime.strptime(month_str, "%Y-%m")
        except ValueError:
            errors.append("Invalid month format. Use YYYY-MM.")

    return amount, errors


@app.route("/budgets")
def budgets():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    # Determine selected month (default: current month)
    month = request.args.get("month", "").strip()
    if month:
        # Validate month format; fall back to current month on bad input
        try:
            month_date = datetime.datetime.strptime(month, "%Y-%m")
        except ValueError:
            month_date = datetime.datetime.today()
            month = month_date.strftime("%Y-%m")
    else:
        month_date = datetime.datetime.today()
        month = month_date.strftime("%Y-%m")

    # Calculate previous and next months for navigation links
    if month_date.month == 1:
        prev_month = f"{month_date.year - 1}-12"
    else:
        prev_month = f"{month_date.year}-{month_date.month - 1:02d}"

    if month_date.month == 12:
        next_month = f"{month_date.year + 1}-01"
    else:
        next_month = f"{month_date.year}-{month_date.month + 1:02d}"

    month_label = month_date.strftime("%B %Y")

    # Fetch budgets for this month and actual spending from expenses
    budget_list = get_budgets_for_month(user_id, month)
    spending = get_month_category_spending(user_id, month)

    # Enrich each budget dict with spent amount, percentage, and status
    for b in budget_list:
        spent = spending.get(b["category"], 0.0)
        b["spent"] = round(spent, 2)
        b["percentage"] = (
            round((spent / b["amount"]) * 100, 1) if b["amount"] > 0 else 0.0
        )
        if b["percentage"] > 100:
            b["status"] = "exceeded"
        elif b["percentage"] > 75:
            b["status"] = "warning"
        else:
            b["status"] = "ok"

    return render_template(
        "budgets.html",
        budgets=budget_list,
        month=month,
        month_label=month_label,
        prev_month=prev_month,
        next_month=next_month,
    )


@app.route("/budgets/add", methods=["GET", "POST"])
def add_budget():
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    default_month = datetime.date.today().strftime("%Y-%m")

    if request.method == "POST":
        category = request.form.get("category", "").strip()
        amount_str = request.form.get("amount", "").strip()
        month = request.form.get("month", "").strip()

        amount, errors = validate_budget_data(category, amount_str, month)

        if errors:
            for error in errors:
                flash(error, "error")
            return render_template(
                "add_budget.html",
                categories=ALLOWED_CATEGORIES,
                category=category,
                amount=amount_str,
                month=month or default_month,
            )

        try:
            create_budget(user_id, category, amount, month)
            flash("Budget created successfully!", "success")
            return redirect(url_for("budgets", month=month))
        except sqlite3.IntegrityError:
            flash("A budget for this category and month already exists.", "error")
            return render_template(
                "add_budget.html",
                categories=ALLOWED_CATEGORIES,
                category=category,
                amount=amount_str,
                month=month,
            )

    return render_template(
        "add_budget.html",
        categories=ALLOWED_CATEGORIES,
        category="",
        amount="",
        month=default_month,
    )


@app.route("/budgets/<int:id>/edit", methods=["GET", "POST"])
def edit_budget(id):
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    budget = get_budget_by_id(id)
    if not budget:
        abort(404)
    if budget["user_id"] != user_id:
        abort(403)

    if request.method == "POST":
        category = request.form.get("category", "").strip()
        amount_str = request.form.get("amount", "").strip()
        month = request.form.get("month", "").strip()

        amount, errors = validate_budget_data(category, amount_str, month)

        if errors:
            for error in errors:
                flash(error, "error")
            return render_template(
                "edit_budget.html",
                budget=budget,
                categories=ALLOWED_CATEGORIES,
                category=category,
                amount=amount_str,
                month=month,
            )

        try:
            update_budget(id, category, amount, month)
            flash("Budget updated successfully!", "success")
            return redirect(url_for("budgets", month=month))
        except sqlite3.IntegrityError:
            flash("A budget for this category and month already exists.", "error")
            return render_template(
                "edit_budget.html",
                budget=budget,
                categories=ALLOWED_CATEGORIES,
                category=category,
                amount=amount_str,
                month=month,
            )

    return render_template(
        "edit_budget.html",
        budget=budget,
        categories=ALLOWED_CATEGORIES,
        category=budget["category"],
        amount=budget["amount"],
        month=budget["month"],
    )


@app.route("/budgets/<int:id>/delete", methods=["GET", "POST"])
def delete_budget_route(id):
    user_id = session.get("user_id")
    if not user_id:
        flash("Please log in to access this page.", "error")
        return redirect(url_for("login"))

    budget = get_budget_by_id(id)
    if not budget:
        abort(404)
    if budget["user_id"] != user_id:
        abort(403)

    if request.method == "POST":
        db_delete_budget(id)
        flash("Budget deleted successfully!", "success")
        return redirect(url_for("budgets", month=budget["month"]))

    return render_template("delete_budget.html", budget=budget)


# ------------------------------------------------------------------ #
# Error Handlers & Testing Routes                                    #
# ------------------------------------------------------------------ #


@app.errorhandler(403)
def forbidden_error(error):
    return render_template("errors/403.html"), 403


@app.errorhandler(404)
def not_found_error(error):
    return render_template("errors/404.html"), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template("errors/500.html"), 500


@app.route("/trigger-500")
def trigger_500():
    if not (app.config.get("TESTING") or app.debug):
        abort(404)
    abort(500)


if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port, debug=True)
